from __future__ import annotations

import hashlib
import sqlite3
import json
import re
import csv
import io
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Body
from fastapi.responses import FileResponse, Response
from typing import List
import base64

from app.config import BACKGROUND_IMAGES_DIR, BADGE_PNG, DATA_BASE, ISSUED_DIR, get_max_csv_upload_bytes, get_max_png_upload_bytes
from app.database import get_database_connection, close_connection
from app.database import (
    add_badge_template, get_badge_template_by_id, get_all_badge_templates,
    update_badge_template, delete_badge_template, get_badge_schema_by_id,
    add_batch_session_item, create_batch_session, get_batch_session,
    get_batch_session_items, list_batch_sessions,
)
from app.batch_issuer import parse_batch_file, preview_batch_file, preview_batch_rows
from app.issuer import issue_baked_badge_from_template, make_search_hash, normalize_email
from app.models import BadgeTemplate, TextOverlay
from app.qr import overlay_qr_on_badge, overlay_text_on_badge
from app.upload_limits import ensure_image_pixels_within_limit, read_upload_limited

router = APIRouter()


def get_db():
    """Dépendance FastAPI fournissant une connexion à la base de données."""
    conn = get_database_connection()
    try:
        yield conn
    finally:
        close_connection(conn)


def _dump_model(model):
    """Conversion en dictionnaire compatible Pydantic v2/v1."""
    if hasattr(model, "model_dump"):
        return model.model_dump()
    return model.dict()


def _template_update_payload(template_id: str, template: dict, **overrides) -> dict:
    """Construit une charge complète en conservant les valeurs actuelles et les surcharges."""
    payload = {
        "id": template_id,
        "name": template["name"],
        "description": template.get("description"),
        "schema_id": template.get("schema_id"),
        "background_image": template.get("background_image"),
        "text_overlays": template.get("text_overlays", []),
        "qr_code_placement": template.get("qr_code_placement", "bottom-right"),
        "qr_code_size": template.get("qr_code_size", 0.22),
        "qr_code_offset_x": template.get("qr_code_offset_x", 0),
        "qr_code_offset_y": template.get("qr_code_offset_y", 0),
        "qr_code_foreground_color": template.get("qr_code_foreground_color", "#000000"),
        "qr_code_background_color": template.get("qr_code_background_color", "#FFFFFF"),
        "qr_code_error_correction": template.get("qr_code_error_correction", "M"),
        "qr_code_border": template.get("qr_code_border", 2),
        "is_active": template.get("is_active", True),
        "created_at": template.get("created_at"),
    }
    payload.update(overrides)
    return payload


def _load_template_background(template: dict) -> bytes:
    """Retourne le PNG de fond du modèle, ou l'image de badge par défaut."""
    background_image = template.get("background_image")
    if isinstance(background_image, str) and background_image.startswith("data:") and ";base64," in background_image:
        _, encoded = background_image.split(";base64,", 1)
        try:
            content = base64.b64decode(encoded, validate=True)
        except Exception:
            raise HTTPException(status_code=400, detail="Le fond du badge doit être un PNG encodé en base64 valide")
        _ensure_png_background(content)
        ensure_image_pixels_within_limit(content, label="Fond de badge")
        return content
    if isinstance(background_image, str) and background_image.strip():
        content = _load_named_background_image(background_image)
        _ensure_png_background(content)
        ensure_image_pixels_within_limit(content, label="Fond de badge")
        return content
    content = BADGE_PNG.read_bytes()
    ensure_image_pixels_within_limit(content, label="Fond de badge")
    return content


def _load_named_background_image(background_image: str) -> bytes:
    """Charge un fond PNG depuis le dossier autorisé sans permettre de traversal."""
    safe_name = str(background_image or "").strip()
    if not safe_name:
        return BADGE_PNG.read_bytes()
    if "/" in safe_name or "\\" in safe_name or safe_name in {".", ".."} or ".." in safe_name.split("."):
        raise HTTPException(status_code=400, detail="Chemin de fond invalide")

    base_dir = BACKGROUND_IMAGES_DIR.resolve()
    candidate = (base_dir / safe_name).resolve()
    if not candidate.is_relative_to(base_dir):
        raise HTTPException(status_code=400, detail="Chemin de fond invalide")
    if not candidate.exists() or not candidate.is_file():
        raise HTTPException(status_code=404, detail="Fond de badge introuvable")
    return candidate.read_bytes()


def _ensure_png_background(content: bytes) -> None:
    """Vérifie que le fond fourni est bien un PNG."""
    if not content or not content.startswith(b"\x89PNG\r\n\x1a\n"):
        raise HTTPException(status_code=400, detail="Le fond du badge doit être un fichier PNG valide")


def _iter_template_certificate_numbers(template_id: str) -> list[str]:
    """Retourne les numéros déjà émis pour un modèle donné."""
    numbers: list[str] = []
    if not ISSUED_DIR.exists():
        return numbers

    for assertion_path in ISSUED_DIR.glob("*.json"):
        try:
            assertion = json.loads(assertion_path.read_text(encoding="utf-8"))
        except Exception:
            continue

        template_meta = assertion.get("badge83_template", {}) if isinstance(assertion.get("badge83_template"), dict) else {}
        if template_meta.get("id") != template_id:
            continue

        field_values = assertion.get("field_values", {}) if isinstance(assertion.get("field_values"), dict) else {}
        certificate_number = str(field_values.get("certificate_number") or "").strip()
        if certificate_number:
            numbers.append(certificate_number)

    return numbers


def _next_certificate_number(existing_numbers: list[str]) -> str:
    """Génère le prochain numéro en conservant si possible le préfixe et le padding."""
    if not existing_numbers:
        return "001"

    parsed: list[tuple[int, str, int]] = []
    for number in existing_numbers:
        match = re.search(r"^(.*?)(\d+)$", number)
        if match:
            prefix, digits = match.groups()
            parsed.append((int(digits), prefix, len(digits)))

    if not parsed:
        return str(len(existing_numbers) + 1).zfill(3)

    highest_value, prefix, width = max(parsed, key=lambda item: item[0])
    return f"{prefix}{str(highest_value + 1).zfill(width)}"


def _certificate_number_exists(template_id: str, certificate_number: str) -> bool:
    normalized = str(certificate_number or "").strip()
    if not normalized:
        return False
    return normalized in _iter_template_certificate_numbers(template_id)


def _get_required_batch_field_ids(template: dict, db: sqlite3.Connection) -> list[str]:
    """Retourne les champs requis du schéma utiles pour l'import groupé."""
    schema_id = template.get("schema_id")
    if not schema_id:
        return []
    schema = get_badge_schema_by_id(db, schema_id)
    fields = (schema or {}).get("fields", [])
    return [
        field.get("id")
        for field in fields
        if field.get("required") and field.get("id") and field.get("id") != "certificate_number"
    ]


def _get_batch_schema_fields(template: dict, db: sqlite3.Connection) -> list[dict]:
    """Retourne les champs du schéma pour l'import groupé avec labels et aliases."""
    schema_id = template.get("schema_id")
    if not schema_id:
        return []
    schema = get_badge_schema_by_id(db, schema_id)
    fields = (schema or {}).get("fields", [])
    return [
        field
        for field in fields
        if field.get("id") and field.get("id") != "certificate_number"
    ]


def _batch_template_column_for_field(field: dict) -> str:
    """Retourne l'en-tête Excel opérateur à utiliser pour un champ de schéma."""
    return str(field.get("label") or field.get("id") or "").strip()


def _build_batch_issue_template_xlsx(*, template: dict, schema_fields: list[dict]) -> bytes:
    """Génère un modèle Excel adapté au schéma du modèle de badge sélectionné."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.comments import Comment
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Le support Excel nécessite la dépendance openpyxl") from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Emission groupee"

    base_headers = ["nom", "email", "reussi"]
    field_headers: list[str] = []
    field_by_header: dict[str, dict] = {}
    for field in schema_fields:
        header = _batch_template_column_for_field(field)
        if not header:
            continue
        normalized = header.strip().lower()
        if normalized in {"nom", "name", "email", "reussi", "réussi", "passed"}:
            continue
        if header not in field_by_header:
            field_headers.append(header)
            field_by_header[header] = field

    headers = [*base_headers, *field_headers]
    worksheet.append(headers)

    sample_row = ["Alice Example", "alice@example.org", "oui"]
    for header in field_headers:
        field = field_by_header[header]
        field_type = str(field.get("field_type") or "text")
        label = str(field.get("label") or header).lower()
        if field_type == "email" or "couriel" in label or "courriel" in label or "email" in label:
            sample_row.append("alice@example.org")
        elif field_type == "date" or "date" in label:
            sample_row.append("2026-05-19")
        elif field_type == "number":
            sample_row.append("1")
        elif "cours" in label or "course" in label or "programme" in label:
            sample_row.append("Formation MODE83")
        else:
            sample_row.append("Exemple")
    worksheet.append(sample_row)

    header_fill = PatternFill("solid", fgColor="E8F1FF")
    required_fill = PatternFill("solid", fgColor="FFE4E6")
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index)
        field = field_by_header.get(header)
        is_required = header in base_headers or bool(field and field.get("required"))
        cell.font = Font(bold=True)
        cell.fill = required_fill if is_required else header_fill
        if field:
            required_text = "obligatoire" if field.get("required") else "optionnel"
            cell.comment = Comment(
                f"Champ du schéma : {field.get('label') or field.get('id')} ({required_text}). "
                "Conservez cet en-tête ou utilisez l'identifiant technique du champ.",
                "Badge83",
            )
        elif header == "reussi":
            cell.comment = Comment("Valeurs acceptées : oui/non, yes/no, true/false, 1/0.", "Badge83")

        worksheet.column_dimensions[cell.column_letter].width = max(14, min(36, len(str(header)) + 6))

    info = workbook.create_sheet("Instructions")
    info.append(["Modèle Badge83", template.get("name") or template.get("id") or "Modèle sélectionné"])
    info.append(["Colonnes minimales", "nom, email, reussi"])
    info.append(["Colonnes du schéma", ", ".join(field_headers) if field_headers else "Aucune"])
    info.append(["Note", "Les colonnes en rouge sont obligatoires. Supprimez la ligne d'exemple avant l'import réel si nécessaire."])
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 96

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _batch_not_issued_reason(row: dict) -> str:
    status = row.get("status")
    errors = [str(error) for error in (row.get("errors") or []) if str(error).strip()]
    if status == "not_passed":
        return "Non admis"
    if status == "duplicate":
        return "Duplicate"
    if any("Email invalide" in error for error in errors):
        return "Email invalide"
    if errors:
        return "; ".join(errors)
    return "Non émis"


def _build_batch_report_csv(*, source_rows: list[dict[str, str]], preview_rows: list[dict], created_badges: list[dict]) -> str:
    created_by_row = {badge.get("row_number"): badge for badge in created_badges}
    source_headers: list[str] = []
    for row in source_rows:
        for key in row.keys():
            if key not in source_headers:
                source_headers.append(key)

    fieldnames = [
        *source_headers,
        "badge83_status",
        "badge83_reason",
        "badge83_png_filename",
        "badge83_assertion_id",
        "badge83_verification_url",
        "badge83_qr_url",
    ]
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for source_row, preview_row in zip(source_rows, preview_rows):
        report_row = dict(source_row)
        created = created_by_row.get(preview_row.get("row_number"))
        if created:
            report_row.update(
                {
                    "badge83_status": "issued",
                    "badge83_reason": "",
                    "badge83_png_filename": created.get("archive_png_filename") or "",
                    "badge83_assertion_id": created.get("assertion_id") or "",
                    "badge83_verification_url": created.get("verification_url") or "",
                    "badge83_qr_url": created.get("qr_url") or "",
                }
            )
        else:
            report_row.update(
                {
                    "badge83_status": "not_issued",
                    "badge83_reason": _batch_not_issued_reason(preview_row),
                    "badge83_png_filename": "",
                    "badge83_assertion_id": "",
                    "badge83_verification_url": "",
                    "badge83_qr_url": "",
                }
            )
        writer.writerow(report_row)
    return output.getvalue()


def _build_batch_report_rows(*, preview_rows: list[dict], created_badges: list[dict]) -> list[dict]:
    """Construit un rapport JSON léger pour le retour immédiat du commit batch."""
    created_by_row = {badge.get("row_number"): badge for badge in created_badges}
    report_rows: list[dict] = []
    for preview_row in preview_rows:
        created = created_by_row.get(preview_row.get("row_number"))
        if created:
            report_rows.append(
                {
                    "row_number": preview_row.get("row_number"),
                    "name": preview_row.get("name"),
                    "email": preview_row.get("email"),
                    "status": "issued",
                    "reason": "",
                    "assertion_id": created.get("assertion_id"),
                    "verification_url": created.get("verification_url"),
                    "qr_url": created.get("qr_url"),
                }
            )
        else:
            report_rows.append(
                {
                    "row_number": preview_row.get("row_number"),
                    "name": preview_row.get("name"),
                    "email": preview_row.get("email"),
                    "status": "not_issued",
                    "reason": _batch_not_issued_reason(preview_row),
                    "source_status": preview_row.get("status"),
                    "assertion_id": None,
                    "verification_url": None,
                    "qr_url": None,
                }
            )
    return report_rows


def _batch_session_status(*, preview: dict, created_count: int) -> str:
    """Détermine le statut métier d'une session d'émission groupée."""
    if preview.get("total_rows", 0) == 0 or not preview.get("can_commit"):
        return "empty"
    if preview.get("errors", 0) or preview.get("skipped_duplicates", 0) or preview.get("skipped_not_passed", 0):
        return "completed_with_errors"
    return "completed" if created_count else "empty"


def _persist_batch_session(
    *,
    db: sqlite3.Connection,
    template_id: str,
    source_filename: str,
    source_content: bytes,
    preview: dict,
    created_badges: list[dict],
) -> str:
    """Enregistre une session batch et toutes ses lignes dans SQLite."""
    session_id = str(uuid4())
    created_by_row = {badge.get("row_number"): badge for badge in created_badges}
    create_batch_session(
        db,
        {
            "id": session_id,
            "template_id": template_id,
            "source_filename": source_filename,
            "source_file_hash": "sha256$" + hashlib.sha256(source_content).hexdigest(),
            "status": _batch_session_status(preview=preview, created_count=len(created_badges)),
            "total_rows": preview.get("total_rows", 0),
            "ready_count": preview.get("ready_rows", 0),
            "issued_count": len(created_badges),
            "error_count": preview.get("errors", 0),
            "duplicate_count": preview.get("skipped_duplicates", 0),
            "not_passed_count": preview.get("skipped_not_passed", 0),
        },
    )

    for row in preview.get("rows", []):
        created = created_by_row.get(row.get("row_number"))
        email = str(row.get("email") or "").strip()
        add_batch_session_item(
            db,
            {
                "session_id": session_id,
                "badge_id": created.get("assertion_id") if created else None,
                "row_number": row.get("row_number"),
                "recipient_name": row.get("name"),
                "recipient_email": email or None,
                "recipient_email_hash": make_search_hash(normalize_email(email)) if email else None,
                "status": "issued" if created else str(row.get("status") or "not_issued"),
                "error_message": "" if created else _batch_not_issued_reason(row),
                "verification_url": created.get("verification_url") if created else None,
            },
        )
    return session_id


def _build_batch_archive_response(*, template_id: str, source_filename: str, source_content: bytes, rows: list[dict[str, str]], preview: dict, created_badges: list[dict], png_entries: list[tuple[str, bytes]]) -> Response:
    """Construit l'ancienne réponse ZIP à partir d'entrées PNG déjà collectées.

    Conservé comme helper simple pour les usages de test ou de compatibilité interne.
    Le flux principal écrit désormais les PNG directement dans le ZIP pour éviter
    de conserver toutes les images en mémoire avant compression.
    """
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    archive_name = f"batch-issue-{timestamp}.zip"
    manifest = {
        "template_id": template_id,
        "source_filename": source_filename,
        "created": len(created_badges),
        "skipped_not_passed": preview["skipped_not_passed"],
        "skipped_duplicates": preview["skipped_duplicates"],
        "errors": preview["errors"],
        "created_badges": created_badges,
        "preview": preview,
    }
    report_csv = _build_batch_report_csv(
        source_rows=rows,
        preview_rows=preview["rows"],
        created_badges=created_badges,
    )
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("source.csv", source_content)
        archive.writestr("rapport_emission.csv", report_csv.encode("utf-8-sig"))
        archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2).encode("utf-8"))
        for filename, png_bytes in png_entries:
            archive.writestr(filename, png_bytes)
    return Response(
        content=buffer.getvalue(),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{archive_name}"',
            "X-Badge83-Created": str(len(created_badges)),
        },
    )


@router.post("/templates", response_model=dict)
def create_badge_template(template: BadgeTemplate, db: sqlite3.Connection = Depends(get_db)):
    """Crée un nouveau modèle de badge."""
    try:
        template_data = _dump_model(template)
        template_id = add_badge_template(db, template_data)
        return {"id": template_id, "message": "Modèle créé avec succès"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/batch-issue/template.xlsx", response_class=FileResponse)
def download_batch_issue_excel_template():
    """Télécharge un modèle Excel prêt à remplir pour l'émission groupée."""
    template_path = DATA_BASE / "sample_batch_issue.xlsx"
    if not template_path.exists():
        raise HTTPException(status_code=404, detail="Modèle Excel d'émission groupée introuvable")
    return FileResponse(
        template_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="badge83-modele-emission-groupee.xlsx",
    )


@router.get("/templates/{template_id}/batch-issue/template.xlsx", response_class=Response)
def download_batch_issue_excel_template_for_template(template_id: str, db: sqlite3.Connection = Depends(get_db)):
    """Télécharge un modèle Excel généré selon le schéma du modèle sélectionné."""
    template = get_badge_template_by_id(db, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="Modèle introuvable")

    content = _build_batch_issue_template_xlsx(
        template=template,
        schema_fields=_get_batch_schema_fields(template, db),
    )
    safe_template_name = re.sub(r"[^A-Za-z0-9_.-]+", "-", str(template.get("name") or template_id)).strip("-") or template_id
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="badge83-modele-emission-groupee-{safe_template_name}.xlsx"'},
    )


@router.get("/templates", response_model=List[dict])
def list_badge_templates(db: sqlite3.Connection = Depends(get_db)):
    """Liste tous les modèles de badge actifs."""
    try:
        templates = get_all_badge_templates(db)
        return templates
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/templates/{template_id}", response_model=dict)
def get_badge_template(template_id: str, db: sqlite3.Connection = Depends(get_db)):
    """Récupère un modèle de badge par identifiant."""
    try:
        template = get_badge_template_by_id(db, template_id)
        if not template:
            raise HTTPException(status_code=404, detail="Modèle introuvable")
        return template
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/templates/{template_id}/preview", response_class=Response)
def preview_badge_template(template_id: str, db: sqlite3.Connection = Depends(get_db)):
    """Génère un aperçu PNG avec textes superposés et placement du QR code configurés."""
    try:
        template = get_badge_template_by_id(db, template_id)
        if not template:
            raise HTTPException(status_code=404, detail="Modèle introuvable")

        png_data = _load_template_background(template)
        png_data = overlay_text_on_badge(png_data, template.get("text_overlays", []))
        png_data = overlay_qr_on_badge(
            png_data,
            "https://mode83.example/verify/qr/preview",
            placement=template.get("qr_code_placement", "bottom-right"),
            size_ratio=float(template.get("qr_code_size", 0.22)),
            offset_x=int(template.get("qr_code_offset_x", 0)),
            offset_y=int(template.get("qr_code_offset_y", 0)),
            foreground_color=template.get("qr_code_foreground_color", "#000000"),
            background_color=template.get("qr_code_background_color", "#FFFFFF"),
            error_correction=template.get("qr_code_error_correction", "M"),
            border=int(template.get("qr_code_border", 2)),
        )
        return Response(content=png_data, media_type="image/png")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/templates/preview-draft", response_class=Response)
def preview_badge_template_draft(payload: dict = Body(...)):
    """Génère un aperçu PNG avant l'enregistrement d'un modèle."""
    try:
        template = dict(payload or {})
        png_data = _load_template_background(template)
        sample_values = {
            "name": "Alice Example",
            "recipient_name": "Alice Example",
            "email": "alice@example.org",
            "recipient_email": "alice@example.org",
            "course_name": "Blockchain Foundations",
            "certificate_number": "BF-2026-001",
            "issue_date": "2026-04-29",
        }
        png_data = overlay_text_on_badge(
            png_data,
            template.get("text_overlays", []),
            field_values=sample_values,
        )
        png_data = overlay_qr_on_badge(
            png_data,
            "https://mode83.example/verify/qr/preview",
            placement=template.get("qr_code_placement", "bottom-right"),
            size_ratio=float(template.get("qr_code_size", 0.22)),
            offset_x=int(template.get("qr_code_offset_x", 0)),
            offset_y=int(template.get("qr_code_offset_y", 0)),
            foreground_color=template.get("qr_code_foreground_color", "#000000"),
            background_color=template.get("qr_code_background_color", "#FFFFFF"),
            error_correction=template.get("qr_code_error_correction", "M"),
            border=int(template.get("qr_code_border", 2)),
        )
        return Response(content=png_data, media_type="image/png")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/templates/{template_id}/issue-baked", response_class=Response)
def issue_badge_from_template(
    template_id: str,
    payload: dict = Body(...),
    db: sqlite3.Connection = Depends(get_db),
):
    """Émet un badge baked en utilisant un modèle du constructeur."""
    try:
        template = get_badge_template_by_id(db, template_id)
        if not template:
            raise HTTPException(status_code=404, detail="Modèle introuvable")

        name = str(payload.get("name") or "").strip()
        email = str(payload.get("email") or "").strip()
        field_values = payload.get("field_values") if isinstance(payload.get("field_values"), dict) else {}

        if not name or not email:
            raise HTTPException(status_code=400, detail="Le nom et l'email sont obligatoires")

        schema_id = template.get("schema_id")
        schema_fields = []
        if schema_id:
            schema = get_badge_schema_by_id(db, schema_id)
            schema_fields = (schema or {}).get("fields", [])
            for field in schema_fields:
                field_id = field.get("id")
                if field.get("required") and field_id and not str(field_values.get(field_id, "")).strip():
                    raise HTTPException(status_code=400, detail=f"Champ obligatoire manquant : {field.get('label') or field_id}")

        has_certificate_number_field = any(field.get("id") == "certificate_number" for field in schema_fields)
        certificate_number = str(field_values.get("certificate_number") or "").strip()
        if has_certificate_number_field or certificate_number:
            if not certificate_number:
                certificate_number = _next_certificate_number(_iter_template_certificate_numbers(template_id))
                field_values["certificate_number"] = certificate_number
            elif _certificate_number_exists(template_id, certificate_number):
                raise HTTPException(
                    status_code=409,
                    detail=f"Le numéro de certificat {certificate_number} existe déjà pour ce modèle",
                )

        png_data = _load_template_background(template)
        result = issue_baked_badge_from_template(
            name=name,
            email=email,
            template=template,
            field_values=field_values,
            png_data=png_data,
        )
        return Response(
            content=result["baked_png_bytes"],
            media_type="image/png",
            headers={
                "Content-Disposition": f'attachment; filename="{result.get("baked_download_filename", f"badge-{result["assertion_id"]}.png")}"',
                "X-Badge83-Assertion-Id": result["assertion_id"],
                **({"X-Badge83-Certificate-Number": certificate_number} if certificate_number else {}),
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/templates/{template_id}/batch-issue/preview", response_model=dict)
async def preview_batch_issue_from_template(
    template_id: str,
    file: UploadFile = File(...),
    db: sqlite3.Connection = Depends(get_db),
):
    """Analyse un fichier CSV d'émission groupée sans créer de badge."""
    try:
        template = get_badge_template_by_id(db, template_id)
        if not template:
            raise HTTPException(status_code=404, detail="Modèle introuvable")
        content = await read_upload_limited(file, get_max_csv_upload_bytes(), label="CSV/XLSX")
        return preview_batch_file(
            template_id=template_id,
            file_bytes=content,
            filename=file.filename or "batch.csv",
            schema_fields=_get_batch_schema_fields(template, db),
        )
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/templates/{template_id}/batch-issue", response_model=dict)
async def commit_batch_issue_from_template(
    template_id: str,
    file: UploadFile = File(...),
    db: sqlite3.Connection = Depends(get_db),
):
    """Émet des badges depuis un fichier CSV/XLSX après validation des lignes."""
    try:
        template = get_badge_template_by_id(db, template_id)
        if not template:
            raise HTTPException(status_code=404, detail="Modèle introuvable")

        content = await read_upload_limited(file, get_max_csv_upload_bytes(), label="CSV/XLSX")
        source_filename = file.filename or "batch.csv"
        rows = parse_batch_file(content, source_filename)
        preview = preview_batch_rows(
            template_id=template_id,
            rows=rows,
            schema_fields=_get_batch_schema_fields(template, db),
        )
        ready_rows = [row for row in preview["rows"] if row["status"] == "ready"]
        existing_numbers = _iter_template_certificate_numbers(template_id)
        created_badges = []
        has_certificate_number_field = any(
            overlay.get("field_id") == "certificate_number"
            for overlay in template.get("text_overlays", [])
            if isinstance(overlay, dict)
        )

        for row in ready_rows:
            field_values = dict(row.get("field_values") or {})
            certificate_number = str(field_values.get("certificate_number") or "").strip()
            if has_certificate_number_field or certificate_number:
                if not certificate_number:
                    certificate_number = _next_certificate_number(existing_numbers)
                    field_values["certificate_number"] = certificate_number
                elif certificate_number in existing_numbers or _certificate_number_exists(template_id, certificate_number):
                    continue
                existing_numbers.append(certificate_number)

            png_data = _load_template_background(template)
            result = issue_baked_badge_from_template(
                name=row["name"],
                email=row["email"],
                template=template,
                field_values=field_values,
                png_data=png_data,
            )
            created_badges.append(
                {
                    "row_number": row["row_number"],
                    "name": row["name"],
                    "email": row["email"],
                    "assertion_id": result["assertion_id"],
                    "certificate_number": certificate_number or None,
                    "png_url": f"/api/badges/{result['assertion_id']}/png",
                    "verification_url": f"/verify/badge/{result['assertion_id']}",
                    "qr_url": f"/verify/qr/{result['assertion_id']}",
                }
            )

        session_id = _persist_batch_session(
            db=db,
            template_id=template_id,
            source_filename=source_filename,
            source_content=content,
            preview=preview,
            created_badges=created_badges,
        )

        return {
            "session_id": session_id,
            "template_id": template_id,
            "issue_policy": preview["issue_policy"],
            "can_commit": preview["can_commit"],
            "message": "Émission groupée terminée" if created_badges else "Aucune ligne prête à émettre",
            "created": len(created_badges),
            "skipped_not_passed": preview["skipped_not_passed"],
            "skipped_duplicates": preview["skipped_duplicates"],
            "errors": preview["errors"],
            "created_badges": created_badges,
            "report_rows": _build_batch_report_rows(
                preview_rows=preview["rows"],
                created_badges=created_badges,
            ),
            "preview": preview,
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/templates/{template_id}/batch-issue/archive", response_class=Response)
async def archive_batch_issue_from_template(
    template_id: str,
    file: UploadFile = File(...),
    db: sqlite3.Connection = Depends(get_db),
):
    """Émet des badges depuis un CSV/XLSX et retourne une archive ZIP avec PNG et rapport."""
    try:
        template = get_badge_template_by_id(db, template_id)
        if not template:
            raise HTTPException(status_code=404, detail="Modèle introuvable")

        content = await read_upload_limited(file, get_max_csv_upload_bytes(), label="CSV/XLSX")
        source_filename = file.filename or "batch.csv"
        rows = parse_batch_file(content, source_filename)
        preview = preview_batch_rows(
            template_id=template_id,
            rows=rows,
            schema_fields=_get_batch_schema_fields(template, db),
        )
        ready_rows = [row for row in preview["rows"] if row["status"] == "ready"]
        existing_numbers = _iter_template_certificate_numbers(template_id)
        created_badges: list[dict] = []
        has_certificate_number_field = any(
            overlay.get("field_id") == "certificate_number"
            for overlay in template.get("text_overlays", [])
            if isinstance(overlay, dict)
        )

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
        archive_name = f"batch-issue-{timestamp}.zip"
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr(f"source{Path(source_filename).suffix.lower() or '.csv'}", content)
            for index, row in enumerate(ready_rows, start=1):
                field_values = dict(row.get("field_values") or {})
                certificate_number = str(field_values.get("certificate_number") or "").strip()
                if has_certificate_number_field or certificate_number:
                    if not certificate_number:
                        certificate_number = _next_certificate_number(existing_numbers)
                        field_values["certificate_number"] = certificate_number
                    elif certificate_number in existing_numbers or _certificate_number_exists(template_id, certificate_number):
                        continue
                    existing_numbers.append(certificate_number)

                png_data = _load_template_background(template)
                result = issue_baked_badge_from_template(
                    name=row["name"],
                    email=row["email"],
                    template=template,
                    field_values=field_values,
                    png_data=png_data,
                )
                archive_png_filename = f"badges/{index}_{result.get('baked_download_filename', f'badge-{result["assertion_id"]}.png')}"
                archive.writestr(archive_png_filename, result["baked_png_bytes"])
                created_badges.append(
                    {
                        "row_number": row["row_number"],
                        "name": row["name"],
                        "email": row["email"],
                        "assertion_id": result["assertion_id"],
                        "certificate_number": certificate_number or None,
                        "archive_png_filename": archive_png_filename,
                        "png_url": f"/api/badges/{result['assertion_id']}/png",
                        "verification_url": f"/verify/badge/{result['assertion_id']}",
                        "qr_url": f"/verify/qr/{result['assertion_id']}",
                    }
                )

            session_id = _persist_batch_session(
                db=db,
                template_id=template_id,
                source_filename=source_filename,
                source_content=content,
                preview=preview,
                created_badges=created_badges,
            )
            for badge in created_badges:
                badge["session_id"] = session_id

            report_csv = _build_batch_report_csv(
                source_rows=rows,
                preview_rows=preview["rows"],
                created_badges=created_badges,
            )
            manifest = {
                "session_id": session_id,
                "template_id": template_id,
                "source_filename": source_filename,
                "created": len(created_badges),
                "skipped_not_passed": preview["skipped_not_passed"],
                "skipped_duplicates": preview["skipped_duplicates"],
                "errors": preview["errors"],
                "created_badges": created_badges,
                "preview": preview,
                "archive_generation": {
                    "mode": "streamed_png_entries",
                    "note": "Les PNG sont écrits dans le ZIP au fil de l'émission afin de réduire la mémoire intermédiaire.",
                },
            }
            archive.writestr("rapport_emission.csv", report_csv.encode("utf-8-sig"))
            archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2).encode("utf-8"))

        response = Response(
            content=buffer.getvalue(),
            media_type="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="{archive_name}"',
                "X-Badge83-Created": str(len(created_badges)),
            },
        )
        response.headers["X-Badge83-Session-Id"] = session_id
        return response
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/batch-sessions", response_model=List[dict])
def list_batch_issue_sessions(db: sqlite3.Connection = Depends(get_db)):
    """Liste les dernières sessions d'émission groupée."""
    try:
        return list_batch_sessions(db)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/batch-sessions/{session_id}", response_model=dict)
def get_batch_issue_session(session_id: str, db: sqlite3.Connection = Depends(get_db)):
    """Retourne le détail d'une session d'émission groupée."""
    try:
        session = get_batch_session(db, session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Session d'émission groupée introuvable")
        return {
            "session": session,
            "items": get_batch_session_items(db, session_id),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/templates/{template_id}/certificate-number", response_model=dict)
def get_template_certificate_number_status(template_id: str, value: str | None = None, db: sqlite3.Connection = Depends(get_db)):
    """Retourne le prochain numéro de certificat et vérifie l'unicité d'une valeur saisie."""
    template = get_badge_template_by_id(db, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="Modèle introuvable")

    existing_numbers = _iter_template_certificate_numbers(template_id)
    checked_value = str(value or "").strip()
    exists = checked_value in existing_numbers if checked_value else False
    return {
        "template_id": template_id,
        "next_certificate_number": _next_certificate_number(existing_numbers),
        "existing_count": len(existing_numbers),
        "checked_value": checked_value or None,
        "exists": exists,
        "warning": (
            f"Le numéro de certificat {checked_value} existe déjà pour ce modèle"
            if exists
            else None
        ),
    }


@router.put("/templates/{template_id}", response_model=dict)
def update_badge_template_endpoint(template_id: str, template: BadgeTemplate, db: sqlite3.Connection = Depends(get_db)):
    """Met à jour un modèle de badge existant."""
    try:
        template_data = _dump_model(template)
        template_data["id"] = template_id
        success = update_badge_template(db, template_data)
        if not success:
            raise HTTPException(status_code=404, detail="Modèle introuvable")
        return {"message": "Modèle mis à jour avec succès"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/templates/{template_id}", response_model=dict)
def delete_badge_template_endpoint(template_id: str, db: sqlite3.Connection = Depends(get_db)):
    """Supprime logiquement un modèle de badge."""
    try:
        success = delete_badge_template(db, template_id)
        if not success:
            raise HTTPException(status_code=404, detail="Modèle introuvable")
        return {"message": "Modèle supprimé avec succès"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/templates/{template_id}/duplicate", response_model=dict)
def duplicate_badge_template(template_id: str, db: sqlite3.Connection = Depends(get_db)):
    """Duplique un modèle pour créer rapidement une variante de programme."""
    try:
        template = get_badge_template_by_id(db, template_id)
        if not template:
            raise HTTPException(status_code=404, detail="Modèle introuvable")

        duplicated = dict(template)
        duplicated["id"] = str(uuid4())
        duplicated["name"] = f"Copie de {template.get('name', 'modèle')}"
        duplicated["created_at"] = None
        duplicated["updated_at"] = None
        duplicated["is_active"] = True
        new_id = add_badge_template(db, duplicated)
        return {"id": new_id, "message": "Modèle dupliqué avec succès"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/templates/{template_id}/background-image", response_model=dict)
async def upload_background_image(template_id: str, file: UploadFile = File(...), db: sqlite3.Connection = Depends(get_db)):
    """Téléverse une image de fond pour un modèle de badge."""
    try:
        template = get_badge_template_by_id(db, template_id)
        if not template:
            raise HTTPException(status_code=404, detail="Modèle introuvable")
        
        # Lit et encode l'image en base64
        content = await read_upload_limited(file, get_max_png_upload_bytes(), label="PNG")
        _ensure_png_background(content)
        ensure_image_pixels_within_limit(content, label="Fond de badge")
        base64_image = base64.b64encode(content).decode('utf-8')
        
        template_data = _template_update_payload(
            template_id,
            template,
            background_image=f"data:{file.content_type or 'image/png'};base64,{base64_image}",
        )
        
        success = update_badge_template(db, template_data)
        if not success:
            raise HTTPException(status_code=404, detail="Modèle introuvable")
        
        return {"message": "Image de fond téléversée avec succès"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/templates/{template_id}/text-overlays", response_model=dict)
def add_text_overlay(template_id: str, overlay: TextOverlay, db: sqlite3.Connection = Depends(get_db)):
    """Ajoute un texte superposé à un modèle de badge."""
    try:
        template = get_badge_template_by_id(db, template_id)
        if not template:
            raise HTTPException(status_code=404, detail="Modèle introuvable")
        
        overlays = template.get("text_overlays", [])
        overlay_data = _dump_model(overlay)
        overlays.append(overlay_data)
        
        template_data = _template_update_payload(template_id, template, text_overlays=overlays)
        
        success = update_badge_template(db, template_data)
        if not success:
            raise HTTPException(status_code=404, detail="Modèle introuvable")
        
        return {"id": overlay_data["id"], "message": "Texte superposé ajouté avec succès"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/templates/{template_id}/text-overlays/{overlay_id}", response_model=dict)
def remove_text_overlay(template_id: str, overlay_id: str, db: sqlite3.Connection = Depends(get_db)):
    """Retire un texte superposé d'un modèle de badge."""
    try:
        template = get_badge_template_by_id(db, template_id)
        if not template:
            raise HTTPException(status_code=404, detail="Modèle introuvable")
        
        overlays = template.get("text_overlays", [])
        original_count = len(overlays)
        overlays = [o for o in overlays if o.get("id") != overlay_id]
        if len(overlays) == original_count:
            raise HTTPException(status_code=404, detail="Texte superposé introuvable")
        
        template_data = _template_update_payload(template_id, template, text_overlays=overlays)
        
        success = update_badge_template(db, template_data)
        if not success:
            raise HTTPException(status_code=404, detail="Modèle introuvable")
        
        return {"message": "Texte superposé retiré avec succès"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))