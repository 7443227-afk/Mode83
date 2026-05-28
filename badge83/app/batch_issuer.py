from __future__ import annotations

import csv
from datetime import date, datetime
import io
import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from app import issuer


RESERVED_COLUMNS = {
    "name",
    "nom",
    "email",
    "programme",
    "program",
    "reussi",
    "réussi",
    "passed",
}

RESERVED_FIELD_VALUE_COLUMNS = RESERVED_COLUMNS | {
    "recipient_name",
    "recipient_email",
}

POSITIVE_VALUES = {"oui", "yes", "true", "1", "reussi", "réussi", "passed", "valide", "validé"}
NEGATIVE_VALUES = {"non", "no", "false", "0", "echoue", "échoué", "failed", "absent"}

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
EMAIL_FIELD_ALIASES = {"email", "e_mail", "mail", "couriel", "courriel", "adresse_email", "adresse_mail"}


@dataclass(frozen=True)
class BatchRow:
    row_number: int
    status: str
    name: str | None
    email: str | None
    field_values: dict[str, str]
    errors: list[str]

    def to_dict(self) -> dict[str, Any]:
        return {
            "row_number": self.row_number,
            "status": self.status,
            "name": self.name,
            "email": self.email,
            "field_values": self.field_values,
            "errors": self.errors,
        }


def normalize_column_name(value: str) -> str:
    """Normalise un nom de colonne CSV pour accepter quelques variantes opérateur."""
    normalized = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    normalized = normalized.replace(" ", "_").replace("-", "_")
    normalized = re.sub(r"_+", "_", normalized)
    return normalized


def _field_aliases(field: dict[str, Any]) -> set[str]:
    """Retourne les noms de colonnes CSV acceptés pour un champ de schéma.

    Les champs créés historiquement peuvent avoir un UUID comme identifiant technique.
    Pour éviter d'imposer ces UUID aux opérateurs, on accepte aussi le libellé lisible
    du champ comme colonne CSV, après la même normalisation que les en-têtes importés.
    """
    aliases: set[str] = set()
    for value in (field.get("id"), field.get("label"), *(field.get("aliases") or [])):
        normalized = normalize_column_name(str(value or ""))
        if normalized:
            aliases.add(normalized)
    return aliases


def _build_schema_field_alias_map(schema_fields: list[dict[str, Any]] | None) -> dict[str, str]:
    alias_map: dict[str, str] = {}
    for field in schema_fields or []:
        field_id = str(field.get("id") or "").strip()
        if not field_id:
            continue
        for alias in _field_aliases(field):
            alias_map[alias] = field_id
    return alias_map


def _is_email_like_schema_field(field: dict[str, Any]) -> bool:
    """Détecte les champs de schéma qui représentent l'email du bénéficiaire."""
    return bool(_field_aliases(field) & EMAIL_FIELD_ALIASES)


def normalize_email_value(value: object) -> str:
    return str(value or "").strip().lower()


def is_passed(value: object) -> bool | None:
    normalized = str(value or "").strip().lower()
    if not normalized:
        return None
    if normalized in POSITIVE_VALUES:
        return True
    if normalized in NEGATIVE_VALUES:
        return False
    return None


def _cell_value_to_text(value: object) -> str:
    """Convertit une valeur Excel/CSV en texte stable pour le flux d'import."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time().isoformat() == "00:00:00" else value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_csv_batch_file(file_bytes: bytes) -> list[dict[str, str]]:
    text = file_bytes.decode("utf-8-sig")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise ValueError("Le fichier CSV doit contenir une ligne d'en-tête")

    rows: list[dict[str, str]] = []
    for raw_row in reader:
        normalized_row: dict[str, str] = {}
        for key, value in raw_row.items():
            normalized_key = normalize_column_name(key or "")
            if normalized_key:
                normalized_row[normalized_key] = _cell_value_to_text(value)
        if any(value for value in normalized_row.values()):
            rows.append(normalized_row)
    return rows


def _parse_xlsx_batch_file(file_bytes: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Le support Excel nécessite la dépendance openpyxl") from exc

    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Le fichier XLSX est invalide ou illisible") from exc

    try:
        worksheet = workbook.active
        header: list[str] | None = None
        rows: list[dict[str, str]] = []

        for excel_row in worksheet.iter_rows(values_only=True):
            values = [_cell_value_to_text(value) for value in excel_row]
            if not any(values):
                continue

            if header is None:
                header = [normalize_column_name(value) for value in values]
                if not any(header):
                    raise ValueError("Le fichier XLSX doit contenir une ligne d'en-tête")
                continue

            normalized_row: dict[str, str] = {}
            for index, key in enumerate(header):
                if not key:
                    continue
                value = values[index] if index < len(values) else ""
                normalized_row[key] = value
            if any(value for value in normalized_row.values()):
                rows.append(normalized_row)

        if header is None:
            raise ValueError("Le fichier XLSX doit contenir une ligne d'en-tête")
        return rows
    finally:
        workbook.close()


def parse_batch_file(file_bytes: bytes, filename: str) -> list[dict[str, str]]:
    """Parse un fichier CSV ou XLSX d'émission groupée."""
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv":
        return _parse_csv_batch_file(file_bytes)
    if suffix == ".xlsx":
        return _parse_xlsx_batch_file(file_bytes)
    if suffix == ".xls":
        raise ValueError("Le format .xls n'est pas supporté : utilisez un fichier .xlsx ou .csv")
    raise ValueError("Format de fichier non supporté : utilisez un fichier .csv ou .xlsx")


def _get_first(row: dict[str, str], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _build_field_values(row: dict[str, str], schema_fields: list[dict[str, Any]] | None = None) -> dict[str, str]:
    alias_map = _build_schema_field_alias_map(schema_fields)
    field_values: dict[str, str] = {}
    for key, value in row.items():
        normalized_value = str(value).strip()
        if key in RESERVED_FIELD_VALUE_COLUMNS or not normalized_value:
            continue
        target_key = alias_map.get(key, key)
        field_values[target_key] = normalized_value
    programme = _get_first(row, "programme", "program")
    if programme:
        field_values.setdefault("programme", programme)
        field_values.setdefault("course_name", programme)
    return field_values


def _required_schema_fields(
    *,
    required_field_ids: list[str] | None = None,
    schema_fields: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    if schema_fields is not None:
        return [field for field in schema_fields if field.get("required") and field.get("id")]
    return [
        {"id": field_id, "label": field_id, "required": True, "aliases": [normalize_column_name(field_id)]}
        for field_id in (required_field_ids or [])
    ]


def _iter_existing_template_email_hashes(template_id: str) -> set[str]:
    existing: set[str] = set()
    issued_dir = issuer.DATA_DIR
    if not issued_dir.exists():
        return existing

    for assertion_path in issued_dir.glob("*.json"):
        try:
            assertion = json.loads(assertion_path.read_text(encoding="utf-8"))
        except Exception:
            continue
        template_meta = assertion.get("badge83_template") if isinstance(assertion.get("badge83_template"), dict) else {}
        if template_meta.get("id") != template_id:
            continue
        search = assertion.get("search") if isinstance(assertion.get("search"), dict) else {}
        email_hash = search.get("email_hash")
        if isinstance(email_hash, str) and email_hash:
            existing.add(email_hash)
    return existing


def preview_batch_rows(
    *,
    template_id: str,
    rows: list[dict[str, str]],
    required_field_ids: list[str] | None = None,
    schema_fields: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Classe les lignes CSV sans émettre de badge."""
    required_fields = _required_schema_fields(required_field_ids=required_field_ids, schema_fields=schema_fields)
    existing_email_hashes = _iter_existing_template_email_hashes(template_id)
    seen_email_hashes: set[str] = set()
    prepared_rows: list[BatchRow] = []

    for index, row in enumerate(rows, start=2):
        errors: list[str] = []
        name = _get_first(row, "nom", "name")
        email = normalize_email_value(_get_first(row, "email"))
        passed_raw = _get_first(row, "reussi", "réussi", "passed")
        passed = is_passed(passed_raw)
        field_values = _build_field_values(row, schema_fields=schema_fields)

        if not name:
            errors.append("Nom manquant")
        if not email or not EMAIL_RE.match(email):
            errors.append("Email invalide")
        if passed is None:
            errors.append("Statut de réussite ambigu ou manquant")

        fallback_alias_map = _build_schema_field_alias_map(required_fields)
        for alias, field_id in fallback_alias_map.items():
            if alias in field_values and field_id not in field_values:
                field_values[field_id] = field_values[alias]

        for field in required_fields:
            field_id = str(field.get("id") or "").strip()
            if field_id in {"name", "nom", "email"}:
                continue
            if not str(field_values.get(field_id, "")).strip() and _is_email_like_schema_field(field) and email:
                field_values[field_id] = email
            if not str(field_values.get(field_id, "")).strip():
                label = str(field.get("label") or field_id).strip()
                errors.append(f"Champ obligatoire manquant : {label}")

        email_hash = issuer.make_search_hash(email) if email else None

        if errors:
            status = "error"
        elif passed is False:
            status = "not_passed"
        elif email_hash and (email_hash in existing_email_hashes or email_hash in seen_email_hashes):
            status = "duplicate"
        else:
            status = "ready"

        if email_hash:
            seen_email_hashes.add(email_hash)

        prepared_rows.append(
            BatchRow(
                row_number=index,
                status=status,
                name=name or None,
                email=email or None,
                field_values=field_values,
                errors=errors,
            )
        )

    return build_batch_summary(template_id=template_id, rows=prepared_rows)


def preview_batch_file(
    *,
    template_id: str,
    file_bytes: bytes,
    filename: str,
    required_field_ids: list[str] | None = None,
    schema_fields: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    rows = parse_batch_file(file_bytes, filename)
    return preview_batch_rows(
        template_id=template_id,
        rows=rows,
        required_field_ids=required_field_ids,
        schema_fields=schema_fields,
    )


def build_batch_summary(*, template_id: str, rows: list[BatchRow]) -> dict[str, Any]:
    counters = {
        "ready": 0,
        "not_passed": 0,
        "duplicate": 0,
        "error": 0,
    }
    for row in rows:
        counters[row.status] = counters.get(row.status, 0) + 1
    can_commit = counters["ready"] > 0

    return {
        "template_id": template_id,
        "issue_policy": "partial_valid_rows_only",
        "total_rows": len(rows),
        "ready_count": counters["ready"],
        "not_passed_count": counters["not_passed"],
        "duplicate_count": counters["duplicate"],
        "error_count": counters["error"],
        "ready_rows": counters["ready"],
        "skipped_not_passed": counters["not_passed"],
        "skipped_duplicates": counters["duplicate"],
        "errors": counters["error"],
        "can_commit": can_commit,
        "message": (
            "L'émission peut être confirmée pour les lignes prêtes"
            if can_commit
            else "Aucune ligne prête à émettre"
        ),
        "rows": [row.to_dict() for row in rows],
    }
