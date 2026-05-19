from __future__ import annotations

import csv
import base64
import io
import json
import zipfile
from pathlib import Path

from fastapi import FastAPI
from fastapi.testclient import TestClient

from app.baker import unbake_badge
from app.routes.badge_constructor import router as badge_constructor_router
from app.routes.badge_constructor import templates as constructor_templates


PNG_SIGNATURE = b"\x89PNG\r\n\x1a\n"
XLSX_SIGNATURE = b"PK\x03\x04"


def make_constructor_client(tmp_path, monkeypatch) -> TestClient:
    db_path = tmp_path / "registry.db"
    monkeypatch.setenv("BADGE83_REGISTRY_DB", str(db_path))

    app = FastAPI()
    app.include_router(badge_constructor_router)
    return TestClient(app)


def test_update_badge_template_via_api_persists_changes(tmp_path, monkeypatch):
    client = make_constructor_client(tmp_path, monkeypatch)

    schema_response = client.post(
        "/badge-constructor/schemas",
        json={
            "name": "Programme test",
            "description": "Schéma utilisé pour tester la mise à jour d'un modèle.",
            "fields": [
                {
                    "id": "course_name",
                    "label": "Nom du cours",
                    "field_type": "text",
                    "required": True,
                    "position": 1,
                }
            ],
        },
    )
    assert schema_response.status_code == 200
    schema_id = schema_response.json()["id"]

    create_response = client.post(
        "/badge-constructor/templates",
        json={
            "name": "Modèle initial",
            "schema_id": schema_id,
            "text_overlays": [
                {
                    "content_type": "static",
                    "static_text": "MODE83",
                    "position_x": 24,
                    "position_y": 24,
                    "font_size": 24,
                    "font_color": "#0f172a",
                }
            ],
            "qr_code_placement": "bottom-right",
            "qr_code_size": 0.22,
        },
    )
    assert create_response.status_code == 200
    template_id = create_response.json()["id"]

    update_response = client.put(
        f"/badge-constructor/templates/{template_id}",
        json={
            "id": template_id,
            "name": "Modèle modifié",
            "schema_id": schema_id,
            "text_overlays": [
                {
                    "content_type": "field",
                    "field_id": "course_name",
                    "position_x": 80,
                    "position_y": 160,
                    "font_size": 28,
                    "font_color": "#123456",
                }
            ],
            "qr_code_placement": "top-left",
            "qr_code_size": 0.35,
        },
    )
    assert update_response.status_code == 200
    assert update_response.json()["message"] == "Modèle mis à jour avec succès"

    get_response = client.get(f"/badge-constructor/templates/{template_id}")
    assert get_response.status_code == 200
    updated = get_response.json()

    assert updated["id"] == template_id
    assert updated["name"] == "Modèle modifié"
    assert updated["schema_id"] == schema_id
    assert updated["qr_code_placement"] == "top-left"
    assert updated["qr_code_size"] == 0.35
    assert updated["text_overlays"] == [
        {
            "id": updated["text_overlays"][0]["id"],
            "content_type": "field",
            "static_text": None,
            "field_id": "course_name",
            "font_family": "Arial",
            "font_size": 28,
            "font_color": "#123456",
            "font_style": [],
            "text_align": "left",
            "position_x": 80,
            "position_y": 160,
            "rotation": 0,
            "opacity": 1.0,
            "outline_width": 0,
            "outline_color": "#FFFFFF",
        }
    ]


def test_preview_draft_returns_png_for_static_and_dynamic_overlays(tmp_path, monkeypatch):
    client = make_constructor_client(tmp_path, monkeypatch)

    response = client.post(
        "/badge-constructor/templates/preview-draft",
        json={
            "name": "Aperçu brouillon",
            "text_overlays": [
                {
                    "content_type": "static",
                    "static_text": "MODE83",
                    "position_x": 24,
                    "position_y": 24,
                    "font_size": 24,
                    "font_color": "#0f172a",
                },
                {
                    "content_type": "field",
                    "field_id": "course_name",
                    "position_x": 48,
                    "position_y": 96,
                    "font_size": 20,
                    "font_color": "#123456",
                },
            ],
            "qr_code_placement": "bottom-right",
            "qr_code_size": 0.22,
        },
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "image/png"
    assert response.content.startswith(PNG_SIGNATURE)


def test_download_batch_issue_excel_template_returns_xlsx(tmp_path, monkeypatch):
    client = make_constructor_client(tmp_path, monkeypatch)

    response = client.get("/badge-constructor/batch-issue/template.xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "badge83-modele-emission-groupee.xlsx" in response.headers["content-disposition"]
    assert response.content.startswith(XLSX_SIGNATURE)


def test_download_batch_issue_excel_template_for_selected_schema_contains_schema_fields(tmp_path, monkeypatch):
    from openpyxl import load_workbook

    client = make_constructor_client(tmp_path, monkeypatch)

    schema_response = client.post(
        "/badge-constructor/schemas",
        json={
            "name": "Programme avec couriel",
            "fields": [
                {
                    "id": "29b19e6e-524e-4f79-9c1d-dec3aa775dbe",
                    "label": "Couriel",
                    "field_type": "email",
                    "required": True,
                    "position": 1,
                },
                {
                    "id": "course_name",
                    "label": "Nom du cours",
                    "field_type": "text",
                    "required": True,
                    "position": 2,
                },
            ],
        },
    )
    assert schema_response.status_code == 200
    schema_id = schema_response.json()["id"]

    template_response = client.post(
        "/badge-constructor/templates",
        json={
            "name": "Modèle Couriel",
            "schema_id": schema_id,
            "text_overlays": [],
            "qr_code_placement": "bottom-right",
            "qr_code_size": 0.22,
        },
    )
    assert template_response.status_code == 200
    template_id = template_response.json()["id"]

    response = client.get(f"/badge-constructor/templates/{template_id}/batch-issue/template.xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert response.content.startswith(XLSX_SIGNATURE)

    workbook = load_workbook(io.BytesIO(response.content), read_only=True)
    try:
        worksheet = workbook["Emission groupee"]
        headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        sample_values = [cell.value for cell in next(worksheet.iter_rows(min_row=2, max_row=2))]
    finally:
        workbook.close()

    assert headers == ["nom", "email", "reussi", "Couriel", "Nom du cours"]
    assert sample_values[3] == "alice@example.org"


def test_preview_draft_rejects_background_path_traversal(tmp_path, monkeypatch, sample_png_bytes):
    client = make_constructor_client(tmp_path, monkeypatch)
    monkeypatch.setattr(constructor_templates, "BADGE_PNG", tmp_path / "badge.png")
    constructor_templates.BADGE_PNG.write_bytes(sample_png_bytes)
    backgrounds_dir = tmp_path / "backgrounds"
    backgrounds_dir.mkdir()
    monkeypatch.setattr(constructor_templates, "BACKGROUND_IMAGES_DIR", backgrounds_dir)

    response = client.post(
        "/badge-constructor/templates/preview-draft",
        json={"name": "Traversal", "background_image": "../issuer_template.json", "text_overlays": []},
    )

    assert response.status_code == 400
    assert "Chemin de fond invalide" in response.json()["detail"]


def test_preview_draft_rejects_missing_named_background(tmp_path, monkeypatch, sample_png_bytes):
    client = make_constructor_client(tmp_path, monkeypatch)
    monkeypatch.setattr(constructor_templates, "BADGE_PNG", tmp_path / "badge.png")
    constructor_templates.BADGE_PNG.write_bytes(sample_png_bytes)
    backgrounds_dir = tmp_path / "backgrounds"
    backgrounds_dir.mkdir()
    monkeypatch.setattr(constructor_templates, "BACKGROUND_IMAGES_DIR", backgrounds_dir)

    response = client.post(
        "/badge-constructor/templates/preview-draft",
        json={"name": "Missing", "background_image": "missing.png", "text_overlays": []},
    )

    assert response.status_code == 400
    assert "Fond de badge introuvable" in response.json()["detail"]


def test_preview_draft_accepts_data_url_background(tmp_path, monkeypatch, sample_png_bytes):
    client = make_constructor_client(tmp_path, monkeypatch)
    encoded_png = base64.b64encode(sample_png_bytes).decode("ascii")

    response = client.post(
        "/badge-constructor/templates/preview-draft",
        json={
            "name": "Data URL",
            "background_image": f"data:image/png;base64,{encoded_png}",
            "text_overlays": [],
        },
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "image/png"
    assert response.content.startswith(PNG_SIGNATURE)


def test_update_unknown_badge_template_returns_404(tmp_path, monkeypatch):
    client = make_constructor_client(tmp_path, monkeypatch)

    response = client.put(
        "/badge-constructor/templates/unknown-template",
        json={
            "name": "Modèle inexistant",
            "text_overlays": [],
            "qr_code_placement": "bottom-right",
            "qr_code_size": 0.22,
        },
    )

    assert response.status_code == 404
    assert response.json()["detail"] == "Modèle introuvable"


def test_updated_badge_template_can_issue_baked_png(
    tmp_path,
    monkeypatch,
    isolated_issuer_env,
):
    client = make_constructor_client(tmp_path, monkeypatch)

    schema_response = client.post(
        "/badge-constructor/schemas",
        json={
            "name": "Programme émission",
            "fields": [
                {
                    "id": "course_name",
                    "label": "Nom du cours",
                    "field_type": "text",
                    "required": True,
                    "position": 1,
                }
            ],
        },
    )
    assert schema_response.status_code == 200
    schema_id = schema_response.json()["id"]

    create_response = client.post(
        "/badge-constructor/templates",
        json={
            "name": "Modèle à modifier",
            "schema_id": schema_id,
            "text_overlays": [
                {
                    "content_type": "static",
                    "static_text": "Ancien texte",
                    "position_x": 16,
                    "position_y": 16,
                    "font_size": 16,
                    "font_color": "#000000",
                }
            ],
            "qr_code_placement": "bottom-right",
            "qr_code_size": 0.22,
        },
    )
    assert create_response.status_code == 200
    template_id = create_response.json()["id"]

    update_response = client.put(
        f"/badge-constructor/templates/{template_id}",
        json={
            "id": template_id,
            "name": "Modèle émission modifié",
            "schema_id": schema_id,
            "text_overlays": [
                {
                    "content_type": "field",
                    "field_id": "course_name",
                    "position_x": 24,
                    "position_y": 64,
                    "font_size": 20,
                    "font_color": "#123456",
                }
            ],
            "qr_code_placement": "top-left",
            "qr_code_size": 0.2,
        },
    )
    assert update_response.status_code == 200

    issue_response = client.post(
        f"/badge-constructor/templates/{template_id}/issue-baked",
        json={
            "name": "Alice Example",
            "email": "alice@example.com",
            "field_values": {"course_name": "Blockchain Foundations"},
        },
    )

    assert issue_response.status_code == 200
    assert issue_response.headers["content-type"] == "image/png"
    assert issue_response.content.startswith(PNG_SIGNATURE)
    assertion_id = issue_response.headers["X-Badge83-Assertion-Id"]
    assert assertion_id

    assertion = unbake_badge(issue_response.content)
    assert assertion["id"] == f"https://tests.mode83.local/assertions/{assertion_id}"
    assert assertion["badge83_template"] == {
        "id": template_id,
        "name": "Modèle émission modifié",
        "schema_id": schema_id,
    }
    assert assertion["field_values"] == {"course_name": "Blockchain Foundations"}

    saved_assertion = isolated_issuer_env["issued_dir"] / f"{assertion_id}.json"
    saved_png = isolated_issuer_env["baked_dir"] / f"{assertion_id}.png"
    assert saved_assertion.exists()
    assert saved_png.exists()


def test_batch_issue_preview_csv_returns_summary(tmp_path, monkeypatch, isolated_issuer_env):
    client = make_constructor_client(tmp_path, monkeypatch)

    schema_response = client.post(
        "/badge-constructor/schemas",
        json={
            "name": "Programme émission groupée",
            "fields": [
                {
                    "id": "course_name",
                    "label": "Nom du cours",
                    "field_type": "text",
                    "required": True,
                    "position": 1,
                }
            ],
        },
    )
    assert schema_response.status_code == 200
    schema_id = schema_response.json()["id"]

    create_response = client.post(
        "/badge-constructor/templates",
        json={
            "name": "Modèle groupé",
            "schema_id": schema_id,
            "text_overlays": [],
            "qr_code_placement": "bottom-right",
            "qr_code_size": 0.22,
        },
    )
    assert create_response.status_code == 200
    template_id = create_response.json()["id"]

    csv_content = (
        "nom,email,programme,reussi\n"
        "Alice Example,alice@example.org,Formation IA,oui\n"
        "Bob Example,bob@example.org,Formation IA,non\n"
        "Invalid Example,invalid,Formation IA,oui\n"
    )
    response = client.post(
        f"/badge-constructor/templates/{template_id}/batch-issue/preview",
        files={"file": ("participants.csv", csv_content, "text/csv")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["total_rows"] == 3
    assert payload["ready_rows"] == 1
    assert payload["skipped_not_passed"] == 1
    assert payload["errors"] == 1


def test_batch_issue_preview_rejects_csv_over_configured_limit(tmp_path, monkeypatch, isolated_issuer_env):
    monkeypatch.setenv("BADGE83_MAX_CSV_UPLOAD_BYTES", "10")
    client = make_constructor_client(tmp_path, monkeypatch)

    create_response = client.post(
        "/badge-constructor/templates",
        json={
            "name": "Modèle limite CSV",
            "text_overlays": [],
            "qr_code_placement": "bottom-right",
            "qr_code_size": 0.22,
        },
    )
    assert create_response.status_code == 200
    template_id = create_response.json()["id"]

    response = client.post(
        f"/badge-constructor/templates/{template_id}/batch-issue/preview",
        files={"file": ("participants.csv", "nom,email,programme,reussi\nAlice,alice@example.org,Formation,oui\n", "text/csv")},
    )

    assert response.status_code == 413
    assert response.json()["detail"] == "CSV/XLSX trop volumineux"


def test_background_upload_rejects_png_over_configured_limit(tmp_path, monkeypatch, isolated_issuer_env, sample_png_bytes):
    monkeypatch.setenv("BADGE83_MAX_PNG_UPLOAD_BYTES", "10")
    client = make_constructor_client(tmp_path, monkeypatch)

    create_response = client.post(
        "/badge-constructor/templates",
        json={
            "name": "Modèle limite PNG",
            "text_overlays": [],
            "qr_code_placement": "bottom-right",
            "qr_code_size": 0.22,
        },
    )
    assert create_response.status_code == 200
    template_id = create_response.json()["id"]

    response = client.post(
        f"/badge-constructor/templates/{template_id}/background-image",
        files={"file": ("background.png", sample_png_bytes, "image/png")},
    )

    assert response.status_code == 413
    assert response.json()["detail"] == "PNG trop volumineux"


def test_batch_issue_commit_csv_creates_only_ready_badges(tmp_path, monkeypatch, isolated_issuer_env):
    client = make_constructor_client(tmp_path, monkeypatch)

    schema_response = client.post(
        "/badge-constructor/schemas",
        json={
            "name": "Programme émission groupée commit",
            "fields": [
                {
                    "id": "course_name",
                    "label": "Nom du cours",
                    "field_type": "text",
                    "required": True,
                    "position": 1,
                }
            ],
        },
    )
    assert schema_response.status_code == 200
    schema_id = schema_response.json()["id"]

    create_response = client.post(
        "/badge-constructor/templates",
        json={
            "name": "Modèle groupé commit",
            "schema_id": schema_id,
            "text_overlays": [
                {
                    "content_type": "field",
                    "field_id": "course_name",
                    "position_x": 24,
                    "position_y": 64,
                    "font_size": 20,
                    "font_color": "#123456",
                }
            ],
            "qr_code_placement": "bottom-right",
            "qr_code_size": 0.22,
        },
    )
    assert create_response.status_code == 200
    template_id = create_response.json()["id"]

    csv_content = (
        "nom,email,programme,reussi\n"
        "Alice Example,alice@example.org,Formation IA,oui\n"
        "Bob Example,bob@example.org,Formation IA,non\n"
        "Invalid Example,invalid,Formation IA,oui\n"
    )
    response = client.post(
        f"/badge-constructor/templates/{template_id}/batch-issue",
        files={"file": ("participants.csv", csv_content, "text/csv")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["session_id"]
    assert payload["issue_policy"] == "partial_valid_rows_only"
    assert payload["can_commit"] is True
    assert payload["created"] == 1
    assert payload["skipped_not_passed"] == 1
    assert payload["errors"] == 1
    assert len(payload["created_badges"]) == 1
    assert [row["status"] for row in payload["report_rows"]] == ["issued", "not_issued", "not_issued"]
    assert payload["report_rows"][1]["reason"] == "Non admis"
    assert payload["report_rows"][2]["reason"] == "Email invalide"

    created = payload["created_badges"][0]
    assertion_id = created["assertion_id"]
    saved_assertion = isolated_issuer_env["issued_dir"] / f"{assertion_id}.json"
    saved_png = isolated_issuer_env["baked_dir"] / f"{assertion_id}.png"
    assert saved_assertion.exists()
    assert saved_png.exists()

    assertion = json.loads(saved_assertion.read_text(encoding="utf-8"))
    assert assertion["admin_recipient"]["email"] == "alice@example.org"
    assert assertion["field_values"]["course_name"] == "Formation IA"

    session_response = client.get(f"/badge-constructor/batch-sessions/{payload['session_id']}")
    assert session_response.status_code == 200
    session_payload = session_response.json()
    assert session_payload["session"]["id"] == payload["session_id"]
    assert session_payload["session"]["template_id"] == template_id
    assert session_payload["session"]["issued_count"] == 1
    assert session_payload["session"]["error_count"] == 1
    assert [item["status"] for item in session_payload["items"]] == ["issued", "not_passed", "error"]
    assert session_payload["items"][0]["badge_id"] == assertion_id

    sessions_response = client.get("/badge-constructor/batch-sessions")
    assert sessions_response.status_code == 200
    assert sessions_response.json()[0]["id"] == payload["session_id"]


def test_batch_issue_commit_without_ready_rows_returns_clear_message(tmp_path, monkeypatch, isolated_issuer_env):
    client = make_constructor_client(tmp_path, monkeypatch)

    create_response = client.post(
        "/badge-constructor/templates",
        json={
            "name": "Modèle groupé sans lignes prêtes",
            "text_overlays": [],
            "qr_code_placement": "bottom-right",
            "qr_code_size": 0.22,
        },
    )
    assert create_response.status_code == 200
    template_id = create_response.json()["id"]

    csv_content = (
        "nom,email,programme,reussi\n"
        "Bob Example,bob@example.org,Formation IA,non\n"
        "Invalid Example,invalid,Formation IA,oui\n"
    )
    response = client.post(
        f"/badge-constructor/templates/{template_id}/batch-issue",
        files={"file": ("participants.csv", csv_content, "text/csv")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["session_id"]
    assert payload["created"] == 0
    assert payload["can_commit"] is False
    assert payload["message"] == "Aucune ligne prête à émettre"
    assert [row["status"] for row in payload["report_rows"]] == ["not_issued", "not_issued"]
    assert payload["report_rows"][0]["reason"] == "Non admis"
    assert payload["report_rows"][1]["reason"] == "Email invalide"

    session_response = client.get(f"/badge-constructor/batch-sessions/{payload['session_id']}")
    assert session_response.status_code == 200
    session_payload = session_response.json()
    assert session_payload["session"]["status"] == "empty"
    assert session_payload["session"]["issued_count"] == 0
    assert [item["status"] for item in session_payload["items"]] == ["not_passed", "error"]



def test_batch_issue_archive_returns_zip_with_png_and_report(tmp_path, monkeypatch, isolated_issuer_env):
    client = make_constructor_client(tmp_path, monkeypatch)

    schema_response = client.post(
        "/badge-constructor/schemas",
        json={
            "name": "Programme archive groupée",
            "fields": [
                {
                    "id": "course_name",
                    "label": "Nom du cours",
                    "field_type": "text",
                    "required": True,
                    "position": 1,
                }
            ],
        },
    )
    assert schema_response.status_code == 200
    schema_id = schema_response.json()["id"]

    create_response = client.post(
        "/badge-constructor/templates",
        json={
            "name": "Modèle archive groupée",
            "schema_id": schema_id,
            "text_overlays": [
                {
                    "content_type": "field",
                    "field_id": "course_name",
                    "position_x": 24,
                    "position_y": 64,
                    "font_size": 20,
                    "font_color": "#123456",
                }
            ],
            "qr_code_placement": "bottom-right",
            "qr_code_size": 0.22,
        },
    )
    assert create_response.status_code == 200
    template_id = create_response.json()["id"]

    csv_content = (
        "nom,email,programme,reussi\n"
        "Alice Archive,alice.archive@example.org,Formation IA,oui\n"
        "Bob Archive,bob.archive@example.org,Formation IA,non\n"
        "Invalid Archive,invalid,Formation IA,oui\n"
    )
    response = client.post(
        f"/badge-constructor/templates/{template_id}/batch-issue/archive",
        files={"file": ("participants.csv", csv_content, "text/csv")},
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/zip"
    assert response.headers["X-Badge83-Created"] == "1"
    assert response.headers["X-Badge83-Session-Id"]

    archive = zipfile.ZipFile(io.BytesIO(response.content))
    names = archive.namelist()
    assert "source.csv" in names
    assert "rapport_emission.csv" in names
    assert "manifest.json" in names
    png_names = [name for name in names if name.startswith("badges/") and name.endswith(".png")]
    assert len(png_names) == 1
    assert archive.read(png_names[0]).startswith(PNG_SIGNATURE)

    manifest = json.loads(archive.read("manifest.json").decode("utf-8"))
    assert manifest["created"] == 1
    assert manifest["skipped_not_passed"] == 1
    assert manifest["errors"] == 1
    assert manifest["created_badges"][0]["session_id"] == response.headers["X-Badge83-Session-Id"]

    report_text = archive.read("rapport_emission.csv").decode("utf-8-sig")
    report_rows = list(csv.DictReader(io.StringIO(report_text)))
    assert [row["badge83_status"] for row in report_rows] == ["issued", "not_issued", "not_issued"]
    assert report_rows[0]["badge83_png_filename"].startswith("badges/")
    assert report_rows[1]["badge83_reason"] == "Non admis"
    assert report_rows[2]["badge83_reason"] == "Email invalide"
